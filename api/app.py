import os
import sys
import time
import threading

# Ensure this file's own directory is importable regardless of the working
# directory Vercel's runtime invokes it from — local `python3 app.py` from
# inside api/ already has this on sys.path, but Vercel's serverless import
# mechanism does not, which causes "No module named 'scraper'" there.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from flask import Flask, request, jsonify, send_file
import csv
import io
from datetime import date
from dataclasses import asdict
from scraper import scrape_offices, scrape_office
from scoring import score_agents, summary_stats
from discovery import discover_agencies
import db

app = Flask(__name__)

_TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")


@app.route("/api/db-test")
def db_test():
    """
    TEMPORARY diagnostic endpoint (June 24, 2026) — confirms the real
    Supabase connection works from the live Vercel deployment, the same
    "test against real bytes, not a mock" principle used throughout
    this project. Safe to remove once Day 1-3 infrastructure is
    confirmed stable; does not expose the connection string itself,
    only a count and a timestamp.
    """
    try:
        conn = db.get_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM offices")
            office_count = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM listing_snapshots")
            snapshot_count = cur.fetchone()[0]
        conn.close()
        return jsonify({
            "connected": True,
            "office_count": office_count,
            "listing_snapshot_count": snapshot_count,
        })
    except Exception as e:
        return jsonify({"connected": False, "error": str(e)}), 500


@app.route("/")
def index():
    # Read the template directly rather than relying on Flask's default
    # template-folder resolution, which has been unreliable across
    # different serverless runtimes (working directory assumptions vary).
    with open(os.path.join(_TEMPLATE_DIR, "index.html"), "r", encoding="utf-8") as f:
        return f.read()


@app.route("/api/bulk-discover-and-seed", methods=["POST"])
def bulk_discover_and_seed():
    """
    Takes a LIST of region names (e.g. ["Brisbane CBD QLD", "Toowoomba
    QLD", "Cairns QLD", ...]) plus a Google Places API key, and for
    each region: runs discover_agencies() (with pagination, up to
    max_pages each), then immediately upserts every discovered agency
    with a real website into the offices table — all in one call,
    rather than round-tripping through /api/discover and
    /api/seed-offices separately for each region by hand.

    Built specifically to add the rest of QLD's regions efficiently
    (June 24, 2026), after manually seeding Dural and Gold Coast one
    at a time via the UI.

    Each region is processed independently — if one region's Places
    API call fails, the rest still proceed; the response reports
    per-region results so failures are visible, not silent.
    """
    data = request.get_json(force=True)
    regions = data.get("regions", [])
    api_key = data.get("apiKey", "").strip()
    max_pages = data.get("maxPagesPerRegion", 5)

    if not regions:
        return jsonify({"error": "No regions provided"}), 400
    if not api_key:
        return jsonify({"error": "No Google Places API key provided"}), 400

    region_results = []
    for region in regions:
        region = region.strip()
        if not region:
            continue

        log_lines = []

        def log(msg):
            log_lines.append(msg)

        try:
            agencies = discover_agencies(region, api_key=api_key, log=log, max_pages=max_pages)
        except Exception as e:
            region_results.append({
                "region": region, "discovered": 0, "added": 0,
                "skipped_no_website": 0, "error": f"Discovery failed: {e}",
            })
            continue

        added_count = 0
        skipped_count = 0
        seed_error = None
        for agency in agencies:
            website = (agency.get("website") or "").strip()
            if not website:
                skipped_count += 1
                continue
            domain = website.replace("https://", "").replace("http://", "").rstrip("/")
            try:
                db.upsert_office(domain, office_name=agency.get("name"), region=region)
                added_count += 1
            except Exception as e:
                seed_error = f"Database error while saving: {e}"
                break  # stop this region's seeding, but other regions still proceed

        region_results.append({
            "region": region,
            "discovered": len(agencies),
            "added": added_count,
            "skipped_no_website": skipped_count,
            "error": seed_error,
        })

    total_added = sum(r["added"] for r in region_results)
    return jsonify({"region_results": region_results, "total_offices_added": total_added})


@app.route("/api/discover", methods=["POST"])
def discover():
    data = request.get_json(force=True)
    area = data.get("area", "").strip()
    api_key = data.get("apiKey", "").strip()
    if not area:
        return jsonify({"error": "No suburb/postcode provided"}), 400
    if not api_key:
        return jsonify({"error": "No Google Places API key provided"}), 400

    log_lines = []

    def log(msg):
        log_lines.append(msg)

    agencies = discover_agencies(area, api_key=api_key, log=log)
    return jsonify({"agencies": agencies, "log": log_lines})


@app.route("/api/seed-offices", methods=["POST"])
def seed_offices():
    """
    Takes a list of discovered agencies (same shape as /api/discover's
    output: [{name, place_id, website}, ...]) and writes them into the
    offices table via upsert_office(), so the daily cron job has
    something to scrape. This is a ONE-TIME (or occasional) action per
    region — distinct from the actual scraping, which happens on a
    schedule via /api/cron-scrape.
    """
    data = request.get_json(force=True)
    agencies = data.get("agencies", [])
    region = data.get("region", "").strip()
    if not agencies:
        return jsonify({"error": "No agencies provided"}), 400

    added = []
    skipped = []
    for agency in agencies:
        website = (agency.get("website") or "").strip()
        if not website:
            skipped.append(agency.get("name", "unknown"))
            continue
        domain = website.replace("https://", "").replace("http://", "").rstrip("/")
        try:
            office_id = db.upsert_office(domain, office_name=agency.get("name"), region=region)
        except Exception as e:
            return jsonify({"error": f"Database error while saving offices: {e}"}), 500
        added.append({"id": office_id, "domain": domain, "name": agency.get("name")})

    return jsonify({"added": added, "skipped_no_website": skipped})


def scrape_office_with_hard_timeout(domain, timeout_seconds, browserless_api_key=None):
    """
    Runs scrape_office() with a HARD wall-clock timeout, using a daemon
    thread — necessary because individual HTTP requests already have
    their own 20-second timeout (REQUEST_TIMEOUT in scraper.py), but a
    single office can have many candidate pages to visit (confirmed
    real case: 17 candidate listing pages on one office), so the TOTAL
    time for one office can exceed the whole cron function's time
    budget even though no single request times out.

    Confirmed real bug this fixes (June 24, 2026): Guardian Realty
    caused an unhandled Vercel function timeout (not a Python
    exception — Vercel kills the process outright), which our
    in-Python time-budget check between offices couldn't catch, since
    that check only runs BEFORE starting an office, not DURING one.

    Returns (listings, error) same as scrape_office(), or
    ([], "Timed out after Ns") if the timeout is hit. The background
    thread is a daemon, so if it's still running when the function
    returns, it won't block the process from completing — it will
    simply be abandoned (its eventual result, if any, is discarded).
    """
    result = {"listings": [], "error": "Unknown — thread did not complete"}

    def run():
        try:
            listings, error = scrape_office(domain, browserless_api_key=browserless_api_key)
            # CONFIRMED REAL BUG (June 24, 2026): scrape_office() returns
            # a list of Listing dataclass INSTANCES, not plain dicts —
            # confirmed by checking scrape_offices() (the existing,
            # working UI endpoint), which converts via
            # [asdict(l) for l in all_listings] before returning. Every
            # new endpoint added today (cron-scrape, and db.py's
            # record_scrape_result) wrongly assumed plain dicts and
            # called .get() on them, causing
            # "AttributeError: 'Listing' object has no attribute 'get'"
            # — a real crash found via Vercel's function logs, NOT a
            # timeout issue despite every symptom looking exactly like
            # one (a generic 500, no specific error visible without
            # checking the logs directly). Converting here means every
            # downstream caller (cron_scrape, db.record_scrape_result)
            # always receives plain dicts, with no need to remember
            # this conversion at each call site.
            result["listings"] = [asdict(listing) for listing in listings]
            result["error"] = error
        except Exception as e:
            result["listings"] = []
            result["error"] = f"Unexpected error during scrape: {e}"

    thread = threading.Thread(target=run, daemon=True)
    thread.start()
    thread.join(timeout=timeout_seconds)

    if thread.is_alive():
        return [], f"Timed out after {timeout_seconds}s (office may be genuinely slow or unreachable)"
    return result["listings"], result["error"]


@app.route("/api/cron-scrape", methods=["GET", "POST"])
def cron_scrape():
    """
    Intended to be called by Vercel Cron (configured in vercel.json),
    but also callable manually via GET/POST for testing. Pulls up to
    `limit` offices due for scraping (never scraped, or last scraped
    over 20 hours ago — see db.get_offices_due_for_scraping), scrapes
    each one using the EXACT SAME scrape_office() logic the interactive
    UI uses, and writes results to listing_snapshots.

    `limit` defaults conservatively low (5) to respect Vercel's
    function timeout — the same lesson learned from the UI's own
    client-side batching. A daily cron run with many due offices will
    simply take several days to work through a large backlog at this
    rate; that's an accepted tradeoff for staying within one function's
    timeout, not a bug.

    TWO REAL PROTECTIONS against one slow/broken office crashing the
    entire batch (found via live testing, June 24, 2026 — a known-slow
    site, Guardian Realty, caused an unhandled 500 that took down
    offices that would otherwise have scraped fine in the same batch):
    1. scrape_office() itself is wrapped in try/except — it's possible
       for it to raise an exception we haven't seen before, not just
       return a clean (listings, error) tuple.
    2. A wall-clock time budget (TIME_BUDGET_SECONDS) is checked before
       starting each office. If we're already close to Vercel's
       function timeout, we stop gracefully and return what we have so
       far, rather than risk getting killed mid-request — any
       not-yet-attempted offices simply remain "due" and get picked up
       by the next cron run.
    """
    TIME_BUDGET_SECONDS = 250  # stay under Vercel's 300s function limit
    PER_OFFICE_TIMEOUT_SECONDS = 60  # hard cap per office; see scrape_office_with_hard_timeout

    limit = int(request.args.get("limit", 5))
    try:
        offices = db.get_offices_due_for_scraping(limit=limit)
    except Exception as e:
        return jsonify({"error": f"Database error while fetching due offices: {e}"}), 500

    start_time = time.monotonic()
    results = []
    for office in offices:
        elapsed = time.monotonic() - start_time
        if elapsed > TIME_BUDGET_SECONDS:
            results.append({
                "domain": office["domain"],
                "listing_count": 0,
                "error": "Skipped — time budget exhausted, will retry on next cron run",
            })
            continue

        try:
            listings, error = scrape_office_with_hard_timeout(
                office["domain"],
                timeout_seconds=PER_OFFICE_TIMEOUT_SECONDS,
                browserless_api_key=os.environ.get("FIRECRAWL_API_KEY"),
            )
        except Exception as e:
            # Should be unreachable now that scrape_office_with_hard_timeout
            # catches its own exceptions internally — kept as a final
            # safety net in case the wrapper itself is ever changed.
            listings, error = [], f"Unexpected error during scrape: {e}"

        platform_detected = None
        if listings:
            platform_detected = listings[0].get("source_adapter")
        try:
            db.record_scrape_result(
                office["id"], listings, platform_detected=platform_detected, error=error
            )
        except Exception as e:
            # A DB write failure for one office shouldn't abort the
            # whole batch — log it in the response and continue, since
            # the next cron run will simply retry this office anyway
            # (its last_scraped_at won't have been updated).
            results.append({
                "domain": office["domain"],
                "listing_count": len(listings),
                "error": f"Scrape OK but DB write failed: {e}",
            })
            continue
        results.append({
            "domain": office["domain"],
            "listing_count": len(listings),
            "error": error,
        })

    return jsonify({"scraped": results})


@app.route("/api/office-status")
def office_status():
    """Simple monitoring view: which offices are working, which are
    failing, and why — the Day 3 monitoring capability from the
    production plan, built on top of db.get_office_status_summary()."""
    try:
        summary = db.get_office_status_summary()
    except Exception as e:
        return jsonify({"error": f"Database error: {e}"}), 500
    # Convert datetime objects to ISO strings so jsonify doesn't choke
    for row in summary:
        for key in ("last_scraped_at", "last_success_at"):
            if row.get(key):
                row[key] = row[key].isoformat()
    return jsonify({"offices": summary})


@app.route("/api/scrape", methods=["POST"])
def scrape():
    data = request.get_json(force=True)
    urls = data.get("urls", [])
    if not urls:
        return jsonify({"error": "No URLs provided"}), 400

    # Optional — only used by GenericFallbackAdapter's tier 4 (LLM
    # extraction), and only when every free tier fails to find a usable
    # address. Sent per-request, never stored server-side, same pattern
    # as the Google Places API key in /api/discover.
    llm_api_key = data.get("llmApiKey", "").strip() or None

    # Optional — only used as a LAST RESORT when every plain-HTTP
    # candidate path finds zero listing URLs at all (confirmed real
    # need: LJ Hooker's HubSpot platform generation, June 2026). Same
    # per-request, never-stored pattern as the other two keys above.
    browserless_api_key = data.get("browserlessApiKey", "").strip() or os.environ.get("FIRECRAWL_API_KEY")

    log_lines = []

    def log(msg):
        log_lines.append(msg)

    result = scrape_offices(
        urls, log=log, llm_api_key=llm_api_key, browserless_api_key=browserless_api_key,
    )
    result["log"] = log_lines

    return jsonify(result)


@app.route("/api/export.csv", methods=["POST"])
def export_csv():
    data = request.get_json(force=True)
    listings = data.get("listings", [])
    if not listings:
        return jsonify({"error": "No data to export — run a scrape first"}), 400

    output = io.StringIO()
    fieldnames = list(listings[0].keys())
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(listings)

    mem = io.BytesIO(output.getvalue().encode("utf-8"))
    return send_file(
        mem,
        mimetype="text/csv",
        as_attachment=True,
        download_name="agentscore_listings.csv",
    )


@app.route("/api/export.xlsx", methods=["POST"])
def export_xlsx():
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed on server"}), 500

    data = request.get_json(force=True)
    listings = data.get("listings", [])
    if not listings:
        return jsonify({"error": "No data to export — run a scrape first"}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listings"

    fieldnames = list(listings[0].keys())
    ws.append(fieldnames)
    for row in listings:
        ws.append([row.get(f, "") for f in fieldnames])

    # Light formatting: bold header, autosize-ish columns
    for col_idx, name in enumerate(fieldnames, start=1):
        ws.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(name) + 2)

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="agentscore_listings.xlsx",
    )


@app.route("/api/export.rankings.xlsx", methods=["POST"])
def export_rankings_xlsx():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed on server"}), 500

    data = request.get_json(force=True)
    listings = data.get("listings", [])
    if not listings:
        return jsonify({"error": "No data to export — run a scrape first"}), 400

    scores = score_agents(listings)
    stats = summary_stats(listings, scores)

    wb = openpyxl.Workbook()

    # --- Sheet 1: Agent Rankings (matches original AgentScore template) ---
    ws = wb.active
    ws.title = "Agent Rankings"

    offices = sorted(set(
        l.get("office_name") for l in listings if l.get("office_name")
    ))
    office_label = ", ".join(offices) if offices else "All offices"

    ws.append([f"AgentScore — {office_label}  |  {date.today().isoformat()}"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=13, name="Arial")
    ws.append([
        "Vendor Variance = (Sold Price - Guide Price) / Guide Price  |  "
        "Listings without both prices excluded  |  Min 3 priced sales to rank  |  "
        "Source: AgentScore live scrape (Ray White Dynamics, Harcourts/Cloudhi)"
    ])
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, name="Arial", color="666666")
    ws.append([])

    headers = ["#", "Agent", "Office", "Total Sales", "Priced Sales", "Avg Variance %", "Accuracy %", "Score", "Confidence", "Notes"]
    header_row = 4
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True, name="Arial")

    for i, s in enumerate(scores, start=1):
        ws.append([
            i,
            s.agent_name,
            s.office_name,
            s.total_sales,
            s.scored_sales,
            round(s.avg_variance_pct, 4),
            round(s.accuracy_pct, 4),
            s.score_band,
            s.confidence,
            s.notes,
        ])

    # Number formatting: percentages as 0.0%, matching xlsx skill conventions
    last_row = header_row + len(scores)
    for row_idx in range(header_row + 1, last_row + 1):
        ws.cell(row=row_idx, column=6).number_format = "0.0%;(0.0%)"
        ws.cell(row=row_idx, column=7).number_format = "0.0%"
        # Medium/mixed confidence rows get a subtle flag so a reader knows
        # to treat them with more caution than high-confidence Ray White rows
        conf_cell = ws.cell(row=row_idx, column=9)
        if conf_cell.value in ("medium", "low", "mixed"):
            conf_cell.font = Font(name="Arial", color="9C6500")

    col_widths = {1: 4, 2: 22, 3: 32, 4: 12, 5: 12, 6: 14, 7: 11, 8: 7, 9: 11, 10: 36}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Use a professional font throughout, per xlsx skill guidance
    for row in ws.iter_rows():
        for cell in row:
            if cell.font is None or cell.font.name != "Arial":
                existing = cell.font
                cell.font = Font(
                    name="Arial",
                    bold=existing.bold if existing else False,
                    italic=existing.italic if existing else False,
                    size=existing.size if existing else 11,
                    color=existing.color if existing else None,
                )

    # --- Sheet 2: Summary Stats ---
    ws2 = wb.create_sheet("Summary Stats")
    ws2.append(["AgentScore — Key Statistics"])
    ws2.cell(row=1, column=1).font = Font(bold=True, size=13, name="Arial")
    ws2.append([])
    for label, value in stats.items():
        ws2.append([label, value])
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 40
    for row in ws2.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", bold=(cell.row == 1))

    # --- Sheet 3: Excluded / Unscored rows for transparency ---
    ws3 = wb.create_sheet("Excluded From Scoring")
    ws3.append([
        "Listings excluded from scoring (missing price, active status, "
        "or agent below the 3-sale minimum). Included here for transparency, "
        "not used in the rankings above."
    ])
    ws3.cell(row=1, column=1).font = Font(italic=True, size=9, name="Arial", color="666666")
    ws3.append([])
    excl_headers = ["Status", "Address", "Agent", "Office", "Guide Price", "Sold Price", "Reason"]
    ws3.append(excl_headers)
    for cell in ws3[3]:
        cell.font = Font(bold=True, name="Arial")

    ranked_agents = {s.agent_name for s in scores}
    for l in listings:
        status = l.get("status", "")
        agent = l.get("agent_name", "")
        guide, sold = l.get("guide_price"), l.get("sold_price")
        source = l.get("source_adapter", "")
        reason = ""
        if status != "Sold":
            reason = "Active listing (not yet sold)"
        elif not guide and not sold:
            reason = "No price published on listing page"
        elif not guide and source == "cloudhi_rex":
            # Confirmed via live inspection: many Harcourts/Cloudhi sold
            # listings genuinely never had a guide price published at all
            # (e.g. "sold without a price" disclaimer) — not a scraping
            # gap, the source data itself only has the final sold figure.
            reason = "No guide price published (sold without disclosed price)"
        elif not guide or not sold:
            reason = "Missing guide or sold price"
        elif agent not in ranked_agents:
            reason = "Agent below 3-sale minimum"
        if reason:
            ws3.append([
                status, l.get("address", ""), agent, l.get("office_name", ""),
                guide, sold, reason,
            ])
    excl_widths = {1: 10, 2: 42, 3: 22, 4: 32, 5: 12, 6: 12, 7: 30}
    for col_idx, width in excl_widths.items():
        ws3.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws3.iter_rows(min_row=4):
        for cell in row:
            cell.font = Font(name="Arial")

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="agentscore_rankings.xlsx",
    )


if __name__ == "__main__":
    app.run(debug=True, port=5050)


@app.route("/api/agent-scores")
def agent_scores_api():
    from flask import jsonify
    from collections import defaultdict
    import datetime

    MIN_SALES = 5

    def grade(v):
        v = abs(v)
        if v <= 3: return "A"
        if v <= 7: return "B"
        if v <= 12: return "C"
        return "D"

    conn = db.get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT agent_name, suburb, guide_price, sold_price
                FROM listing_snapshots
                WHERE agent_name IS NOT NULL AND agent_name != ''
                  AND guide_price ~ '^[0-9]+$'
                  AND sold_price ~ '^[0-9]+$'
                  AND guide_price::numeric > 50000
                  AND sold_price::numeric > 50000
                  AND (sold_price::numeric / guide_price::numeric) BETWEEN 0.5 AND 2.0
            """)
            rows = cur.fetchall()
            cols = [d[0] for d in cur.description]
    finally:
        conn.close()

    agent_map = defaultdict(lambda: {"variances": [], "suburbs": set()})
    for row in rows:
        r = dict(zip(cols, row))
        name = (r.get("agent_name") or "").strip()
        if not name: continue
        guide = int(r["guide_price"])
        sold = int(r["sold_price"])
        variance = (sold - guide) / guide * 100
        agent_map[name]["variances"].append(variance)
        if r.get("suburb"):
            agent_map[name]["suburbs"].add(r["suburb"])

    scored = []
    for name, d in agent_map.items():
        n = len(d["variances"])
        if n < MIN_SALES: continue
        avg = sum(d["variances"]) / n
        scored.append({
            "name": name,
            "sales": n,
            "avg_variance": round(avg, 1),
            "grade": grade(avg),
            "suburbs": sorted(d["suburbs"]),
        })

    scored.sort(key=lambda x: abs(x["avg_variance"]))
    for i, r in enumerate(scored):
        r["rank"] = i + 1

    resp = jsonify({
        "agents": scored,
        "total": len(scored),
        "generated_at": datetime.datetime.utcnow().isoformat() + "Z",
    })
    resp.headers["Cache-Control"] = "public, max-age=3600"
    return resp


@app.route("/scores")
def scores_page():
    from flask import render_template
    return render_template("scores.html")


@app.route("/api/agent-detail")
def agent_detail_api():
    from flask import jsonify, request
    import datetime
    agent_name = request.args.get("name", "").strip()
    if not agent_name:
        return jsonify({"error": "name parameter required"}), 400
    def grade(v):
        v = abs(v)
        if v <= 3: return "A"
        if v <= 7: return "B"
        if v <= 12: return "C"
        return "D"
    conn = db.get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT ls.address, ls.suburb, ls.guide_price, ls.sold_price,
                       ls.status, ls.date_listed, ls.sold_date, ls.days_on_market,
                       ls.listing_url, o.office_name, o.domain
                FROM listing_snapshots ls
                LEFT JOIN offices o ON o.id = ls.office_id
                WHERE ls.agent_name = %s
                ORDER BY ls.scraped_at DESC
            """, (agent_name,))
            rows = cur.fetchall()
            cols = [d[0] for d in cur.description]
    finally:
        conn.close()
    listings = []
    variances = []
    for row in rows:
        r = dict(zip(cols, row))
        guide = int(r["guide_price"]) if r["guide_price"] and str(r["guide_price"]).isdigit() else None
        sold = int(r["sold_price"]) if r["sold_price"] and str(r["sold_price"]).isdigit() else None
        variance = None
        if guide and sold and guide > 50000 and sold > 50000:
            ratio = sold / guide
            if 0.5 <= ratio <= 2.0:
                variance = round((sold - guide) / guide * 100, 1)
                variances.append(variance)
        listings.append({"address": r["address"], "suburb": r["suburb"],
            "guide_price": guide, "sold_price": sold, "status": r["status"],
            "date_listed": r["date_listed"] or "", "sold_date": r["sold_date"] or "",
            "days_on_market": r["days_on_market"] or "", "listing_url": r["listing_url"],
            "office_name": r["office_name"] or "", "domain": r["domain"] or "",
            "variance": variance})
    avg_variance = round(sum(variances)/len(variances), 1) if variances else None
    resp = jsonify({"agent_name": agent_name, "total_listings": len(listings),
        "scored_listings": len(variances), "avg_variance": avg_variance,
        "grade": grade(avg_variance) if avg_variance is not None else None,
        "listings": listings,
        "generated_at": datetime.datetime.utcnow().isoformat() + "Z"})
    resp.headers["Cache-Control"] = "public, max-age=300"
    return resp


@app.route("/agent/<path:agent_name>")
def agent_page(agent_name):
    from flask import render_template
    return render_template("agent.html", agent_name=agent_name)
